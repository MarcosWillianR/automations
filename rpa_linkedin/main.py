from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait as Wait
from selenium_utils import SeleniumUtils
from openpyxl.styles import Font, PatternFill
import pandas as pd
import openpyxl
import queue
import concurrent.futures
import time

proxies = [
    { 'host': '', 'port': '', 'user': '', 'pass': '' }
]

class LinkedinRPA:
    def __init__(self):
        self.__profiles = self.__get_profiles()
        self.__response = []
        self.__utils = SeleniumUtils()
        self.__options = self.__configure_chrome_options()

        self.proxies_queue = queue.Queue()
        for proxy in proxies:
            self.proxies_queue.put(proxy)
    
    def __get_profiles(self):
        profiles = []
        workbook = openpyxl.load_workbook("links.xlsx")
        try:
            sheet = workbook["Planilha1"]
            for row in sheet.iter_rows(values_only=True):
                profiles.append(row[0])
        finally:
            workbook.close()
        return profiles

    def __get_next_proxy(self):
        try:
            proxy = self.proxies_queue.get_nowait()
        except queue.Empty:
            for proxy in proxies:
                self.proxies_queue.put(proxy)
            proxy = self.proxies_queue.get_nowait()
        return proxy

    def __configure_chrome_options(self):
        options = webdriver.ChromeOptions()
        options.add_argument('--disable-infobars')
        options.add_argument('--disable-browser-side-navigation')
        options.add_argument('--disable-notifications')
        options.add_argument('--disable-popup-blocking')
        options.add_argument('--js-flags="--expose-gc"')
        options.add_argument('--enable-precise-memory-info')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--disable-software-rasterizer')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--log-level=3')
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        options.add_argument(f'user-agent={user_agent}')
        return options

    def __get_chrome_driver(self, proxy=None):
        if proxy:
            seleniumwire_options = {
                'proxy': {
                    'http': f'http://{proxy["user"]}:{proxy["pass"]}@{proxy["host"]}:{proxy["port"]}',
                    'https': f'https://{proxy["user"]}:{proxy["pass"]}@{proxy["host"]}:{proxy["port"]}',
                }
            }
        else:
            seleniumwire_options = {}

        driver = webdriver.Chrome(options=self.__options, seleniumwire_options=seleniumwire_options)
        return driver

    def get_profile_data(self, profile_url):
        profile_dict = {
            "nome": "",
            "empresa": ""
        }

        for proxy in proxies:
            current_proxy = self.__get_next_proxy()
            driver = self.__get_chrome_driver(current_proxy)

            try:
                driver.get(profile_url)
                time.sleep(2)

                if (
                    driver.title == 'Cadastre-se | LinkedIn' or 
                    driver.title == '' or 
                    driver.title == 'LinkedIn: entre ou cadastre-se' or
                    driver.title == 'Verificação de segurança | LinkedIn'
                ):
                    continue
                
                if driver.title == 'Perfil não encontrado | LinkedIn':
                    print(profile_url)

                popup_close_btn = self.__utils.get_element(driver, By.XPATH, '//*[@id="public_profile_contextual-sign-in"]/div/section/button')
                if not isinstance(popup_close_btn, bool):
                    try:
                        popup_close_btn.click()
                    except:
                        pass

                profile_name = self.__utils.get_element(driver, By.XPATH, '//*[@id="main-content"]/section[1]/div/section/section[1]/div/div[2]/div[1]/button/h1')
                if not isinstance(profile_name, bool):
                    profile_dict["nome"] = profile_name.text

                company_name = self.__utils.get_element(driver, By.XPATH, '//*[@id="main-content"]/section[1]/div/section/section[1]/div/div[2]/div[2]/div/div[1]/a/span')
                if isinstance(company_name, bool):
                    company_name = self.__utils.get_element(driver, By.XPATH, '//*[@id="main-content"]/section[1]/div/section/section[1]/div/div[2]/div[2]/div/div[1]/div/span')
                    if not isinstance(company_name, bool):
                        profile_dict["empresa"] = company_name.text
                else:
                    profile_dict["empresa"] = company_name.text
                break
            except Exception as e:
                print(e)
                pass

        if profile_dict["nome"] and profile_dict["empresa"]:
            self.__response.append(profile_dict)

    def startX(self):
        for profile in self.__profiles:
            self.get_profile_data(profile)

        df = pd.DataFrame(self.__response)
        df.to_excel("dados.xlsx", index=False)

linkedin_robot = LinkedinRPA()
linkedin_robot.startX()